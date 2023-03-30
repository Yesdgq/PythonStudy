#!/usr/bin/python3

import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import datetime

fileName = ''
path = os.getcwd()
files=os.listdir(path)
for file in files:
 if '.xlsx' in file or '.xls' in file:
    fileName = file
    print("fileName: "+fileName)
    break

filePath = os.path.abspath(os.path.join(os.getcwd()))
filePath = filePath + '/' + fileName
print(filePath)

myworkbook = load_workbook(filePath) 
print(myworkbook.sheetnames) 
worksheet = myworkbook['房建（在施）']

# sheet表的最大行、最大列
rows = worksheet.max_row #行数
columns = worksheet.max_column #列数

worksheet['ao2'].value = '核查结果'
worksheet['ao2'].alignment = Alignment(horizontal='center', vertical='center')
# 设置行高
# worksheet.row_dimensions[1].height=30
# 设置列宽
worksheet.column_dimensions['Ao'].width=25

cell = worksheet['Q2']
if cell.value != '注册日期':
    print('文件格式错误(1001)')
    os._exit(0)

cell_a = worksheet['S2']
if cell_a.value != '监督计划':
    print('文件格式错误(1002)')
    os._exit(0)

cell_b = worksheet['T2']
if cell_b.value != '首次监督工作会':
    print('文件格式错误(1003)')
    os._exit(0)

cell_c = worksheet['V2']
if cell_c.value != '历次检查日期':
    print('文件格式错误(1004)')
    os._exit(0)

cell_d = worksheet['R2']
if cell_d.value != '风险等级':
    print('文件格式错误(1005)')
    os._exit(0)

cell_e = worksheet['X2']
if cell_e.value != '下次检查月份':
    print('文件格式错误(1006)')
    os._exit(0)

cell_f = worksheet['W2']
if cell_f.value != '已查次数':
    print('文件格式错误(1007)')
    os._exit(0)

for rowss in worksheet.iter_rows(min_row=3, max_row=rows, min_col=23, max_col=24):
    for celll in rowss:
        celll.value = ''
        celll.fill = PatternFill(patternType='solid',fgColor='FFFFFF')

# 遍历注册日期
for i in range(3, rows+1):
    currCell = worksheet.cell(row=i,column=cell.column)
    timeStr = str(currCell.value).strip()
    if timeStr == 'None' or timeStr == None or len(timeStr) == 0:
        continue
    # print('坐标：' + str(i) + ',' + str(cell.column)  + '  ' + timeStr)

    try:
     middle_time = datetime.datetime.strptime(timeStr,'%Y.%m.%d')
    #  print(middle) #输出：2022-02-28 00:00:00
    except ValueError:
     try:
      middle_time = datetime.datetime.strptime(timeStr,'%Y/%m/%d %H:%M:%S')
     except ValueError:
      try:
       middle_time = datetime.datetime.strptime(timeStr,'%Y/%m/%d')
      except ValueError:  
       try:
        middle_time = datetime.datetime.strptime(timeStr,'%Y-%m-%d %H:%M:%S')
       except ValueError:  
        try:
         middle_time = datetime.datetime.strptime(timeStr,'%Y-%m-%d')
        except ValueError:  
         print('注册日期格式不正确(10004)')

    end_date = datetime.datetime.strftime(middle_time, "%Y-%m-%d")
    print(end_date) #输出：2022-02-28
    st_timeStamp = int(time.mktime(middle_time.timetuple()))
    # print(middle_time)
    # print(st_timeStamp)

    # 当前时间戳
    t = int(time.time())
    time_difference = t - st_timeStamp
    cell_curr = worksheet.cell(row=i,column=worksheet['Ao3'].column)
    # cell对齐方式
    cell_curr.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    isFinsh_tash_A = False
    # 1、是否编制完成监督计划
    if time_difference > 3600 * 24 * 3:
        cell_a_sub = worksheet.cell(row=i,column=cell_a.column)
        cell_a_sub_value = cell_a_sub.value
        if cell_a_sub_value == 'None' or cell_a_sub_value == None or len(cell_a_sub_value) == 0: #值为空
            cell_curr.value = '未完成监督计划编制'

            cell_curr.fill = PatternFill(patternType='solid',fgColor='FFE599')
            cell_curr.font = Font(name="Arial", size=11, color="FF2F92", underline="none")

            isFinsh_tash_A = False
        else:
            is_contain = '已编制' in cell_a_sub_value
            if is_contain != True:
                cell_curr.value = '未完成监督计划编制'

                cell_curr.fill = PatternFill(patternType='solid',fgColor='FFE599')
                cell_curr.font = Font(name="Arial", size=11, color="FF2F92", underline="none")

                isFinsh_tash_A = False
            else:
                isFinsh_tash_A = True

    # 2、是否召开首次监督会
    isFinsh_tash_B = False
    if time_difference > 3600 * 24 * 7:
        cell_b_sub = worksheet.cell(row=i,column=cell_b.column)
        cell_b_sub_value = cell_b_sub.value
        # print(cell_b_sub_value)
        if cell_b_sub_value == 'None' or cell_b_sub_value == None or len(cell_b_sub_value) == 0: #值为空 
            prompt = cell_curr.value
            if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                prompt = '未召开首次监督会'
            else:
                prompt = prompt + '\n未召开首次监督会'

            cell_curr.value = prompt

            cell_curr.fill = PatternFill(patternType='solid',fgColor='B9E0A5')
            # cell字体格式设置
            cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")
            
            isFinsh_tash_B = False
        else:
            isFinsh_tash_B = True 

        

    # 3、下次检查时间
    cell_c_sub = worksheet.cell(row=i,column=cell_c.column)
    cell_c_sub_value = cell_c_sub.value

    if isFinsh_tash_A == True and isFinsh_tash_B == True:
        if cell_c_sub_value == 'None' or cell_c_sub_value == None or len(cell_c_sub_value) == 0: #值为空 
        
            prompt = cell_curr.value
            if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                prompt = '未启动首次检查'
            else:
                prompt = prompt + '\n未启动首次检查'

            cell_curr.value = prompt

            cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
            # cell字体格式设置
            cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

        else:
            count_reply = cell_c_sub_value.count('待回复')
            prompt = cell_curr.value

            # 整改报告待回收提示
            if count_reply > 0:
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '有{}项整改报告待回收'.format(count_reply)
                else:
                    prompt = prompt + '\n有{}项整改报告待回收'.format(count_reply)

                cell_curr.value = prompt
            
            dateArr = cell_c_sub_value.split('、')
            
            # 已查次数
            cell_f_sub = worksheet.cell(row=i,column=cell_f.column)
            cell_f_sub.value = len(dateArr)

            lastDate = dateArr[-1]
            lastDate = lastDate.replace('（待回复）', '').strip()
            print('最后一次检查日期:{}'.format(lastDate))

            try:
                middle_time = datetime.datetime.strptime(lastDate,'%Y.%m.%d')
                #  print(middle) #输出：2022-02-28 00:00:00
            except ValueError:
                try:
                    middle_time = datetime.datetime.strptime(lastDate,'%Y/%m/%d')
                except ValueError:  
                    try:
                        middle_time = datetime.datetime.strptime(lastDate,'%Y-%m-%d')
                    except ValueError:  
                        print('历次检查日期格式不正确(10008)')

            # 最后一次检查的时间戳
            lastDate_timeStamp = int(time.mktime(middle_time.timetuple()))
            nextDate_timeStamp = lastDate_timeStamp
            # print(lastDate_timeStamp)

            cell_d_sub = worksheet.cell(row=i,column=cell_d.column)
            risk_level = cell_d_sub.value
            
            if risk_level == 'None' or risk_level == None or len(risk_level) == 0: #值为空 
                prompt = cell_curr.value
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '缺少风险等级'
                else:
                    prompt = prompt + '\n缺少风险等级'
                cell_curr.value = prompt

                cell_curr.fill = PatternFill(patternType='solid',fgColor='FF99FF')
                # cell字体格式设置
                cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('低风险' in risk_level) == True: #只需检查一次
                print('低风险')

                prompt = cell_curr.value
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '已完成检查'
                else:
                    prompt = prompt + '\n已完成检查'
                cell_curr.value = prompt


                cell_curr.fill = PatternFill(patternType='solid',fgColor='B9E0A5')
                cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = ''
                

            elif ('一般风险' in risk_level) == True:
                print('一般风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 * 3

                if len(dateArr) < 3:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '检查次数不足3次'
                    else:
                        prompt = prompt + '\n检查次数不足3次'

                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F8CECC')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F8CECC')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('较大风险' in risk_level) == True:
                print('较大风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 * 3

                if len(dateArr) < 4:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '检查次数不足4次'
                    else:
                        prompt = prompt + '\n检查次数不足4次'

                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('重大风险' in risk_level) == True:
                print('重大风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='EA6B66')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")


    # 当前时间戳
    t = int(time.time())
    time_difference = t - st_timeStamp
    

    #cell边框设置
    pink = "00FF00FF"
    black = '333333'
    green = "00008000"
    thin = Side(border_style="thin", color=black)
    double = Side(border_style="double", color=green)
    cell_curr.border = Border(top=thin, left=double, right=thin, bottom=thin)










worksheet = myworkbook['市政（在施）']
# sheet表的最大行、最大列
rows = worksheet.max_row #行数
columns = worksheet.max_column #列数

cell = worksheet['Z2']
cell.value = '核查结果'
cell.alignment = Alignment(horizontal='center', vertical='center')


# 设置行高
# worksheet.row_dimensions[1].height=30
# 设置列宽
letter=get_column_letter(columns)
worksheet.column_dimensions['Z'].width=25

cell = worksheet['P2']
if cell.value != '注册日期':
    print('文件格式错误(1008)')
    os._exit(0)

cell_a = worksheet['R2']
if cell_a.value != '监督计划':
    print('文件格式错误(1009)')
    os._exit(0)

cell_b = worksheet['S2']
if cell_b.value != '首次监督工作会':
    print('文件格式错误(1010)')
    os._exit(0)

cell_c = worksheet['U2']
if cell_c.value != '历次检查日期':
    print('文件格式错误(1011)')
    os._exit(0)

cell_d = worksheet['Q2']
if cell_d.value != '风险等级':
    print('文件格式错误(1012)')
    os._exit(0)

cell_e = worksheet['W2']
if cell_e.value != '下次检查月份':
    print('文件格式错误(1013)')
    os._exit(0)

cell_f = worksheet['V2']
if cell_f.value != '已查次数':
    print('文件格式错误(1014)')
    os._exit(0)

for rowss in worksheet.iter_rows(min_row=3, max_row=rows, min_col=22, max_col=23):
    for celll in rowss:
        celll.value = ''
        celll.fill = PatternFill(patternType='solid',fgColor='FFFFFF')


# 遍历注册日期
for i in range(3, rows+1):
    currCell = worksheet.cell(row=i,column=cell.column)
    timeStr = str(currCell.value).strip()
    if timeStr == 'None' or timeStr == None or len(timeStr) == 0:
        continue
    # print('坐标：' + str(i) + ',' + str(cell.column)  + '  ' + timeStr)

    try:
     middle_time = datetime.datetime.strptime(timeStr,'%Y.%m.%d')
    #  print(middle) #输出：2022-02-28 00:00:00
    except ValueError:
     try:
      middle_time = datetime.datetime.strptime(timeStr,'%Y/%m/%d %H:%M:%S')
     except ValueError:
      try:
       middle_time = datetime.datetime.strptime(timeStr,'%Y/%m/%d')
      except ValueError:  
       try:
        middle_time = datetime.datetime.strptime(timeStr,'%Y-%m-%d %H:%M:%S')
       except ValueError:  
        try:
         middle_time = datetime.datetime.strptime(timeStr,'%Y-%m-%d')
        except ValueError:  
         print('注册日期格式不正确(10004)')

    # end_date = datetime.datetime.strftime(middle_time, "%Y-%m-%d")
    # print(end_date) #输出：2022-02-28
    st_timeStamp = int(time.mktime(middle_time.timetuple()))
    # print(middle_time)
    # print(st_timeStamp)

    # 当前时间戳
    t = int(time.time())
    time_difference = t - st_timeStamp
    cell_curr = worksheet.cell(row=i,column=worksheet['Z3'].column)
    # cell对齐方式
    cell_curr.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    isFinsh_tash_A = False
    # 1、是否编制完成监督计划
    if time_difference > 3600 * 24 * 3:
        cell_a_sub = worksheet.cell(row=i,column=cell_a.column)
        cell_a_sub_value = cell_a_sub.value
        if cell_a_sub_value == 'None' or cell_a_sub_value == None or len(cell_a_sub_value) == 0: #值为空
            cell_curr.value = '未完成监督计划编制'

            cell_curr.fill = PatternFill(patternType='solid',fgColor='FFE599')
            cell_curr.font = Font(name="Arial", size=11, color="FF2F92", underline="none")

            isFinsh_tash_A = False
        else:
            is_contain = '已编制' in cell_a_sub_value
            if is_contain != True:
                cell_curr.value = '未完成监督计划编制'

                cell_curr.fill = PatternFill(patternType='solid',fgColor='FFE599')
                cell_curr.font = Font(name="Arial", size=11, color="FF2F92", underline="none")

                isFinsh_tash_A = False
            else:
                isFinsh_tash_A = True

    # 2、是否召开首次监督会
    isFinsh_tash_B = False
    if time_difference > 3600 * 24 * 7:
        cell_b_sub = worksheet.cell(row=i,column=cell_b.column)
        cell_b_sub_value = cell_b_sub.value
        # print(cell_b_sub_value)
        if cell_b_sub_value == 'None' or cell_b_sub_value == None or len(cell_b_sub_value) == 0: #值为空 
            prompt = cell_curr.value
            if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                prompt = '未召开首次监督会'
            else:
                prompt = prompt + '\n未召开首次监督会'

            cell_curr.value = prompt

            cell_curr.fill = PatternFill(patternType='solid',fgColor='B9E0A5')
            # cell字体格式设置
            cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")
            
            isFinsh_tash_B = False
        else:
            isFinsh_tash_B = True 

        

    # 3、下次检查时间
    cell_c_sub = worksheet.cell(row=i,column=cell_c.column)
    cell_c_sub_value = cell_c_sub.value

    if isFinsh_tash_A == True and isFinsh_tash_B == True:
        if cell_c_sub_value == 'None' or cell_c_sub_value == None or len(cell_c_sub_value) == 0: #值为空 
        
            prompt = cell_curr.value
            if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                prompt = '未启动首次检查'
            else:
                prompt = prompt + '\n未启动首次检查'

            cell_curr.value = prompt

            cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
            # cell字体格式设置
            cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

        else:
            count_reply = cell_c_sub_value.count('待回复')
            prompt = cell_curr.value

            # 整改报告待回收提示
            if count_reply > 0:
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '有{}项整改报告待回收'.format(count_reply)
                else:
                    prompt = prompt + '\n有{}项整改报告待回收'.format(count_reply)

                cell_curr.value = prompt
            
            dateArr = cell_c_sub_value.split('、')
            
            # 已查次数
            cell_f_sub = worksheet.cell(row=i,column=cell_f.column)
            cell_f_sub.value = len(dateArr)

            lastDate = dateArr[-1]
            lastDate = lastDate.replace('（待回复）', '').strip()
            print('最后一次检查日期:{}'.format(lastDate))

            try:
                middle_time = datetime.datetime.strptime(lastDate,'%Y.%m.%d')
                #  print(middle) #输出：2022-02-28 00:00:00
            except ValueError:
                try:
                    middle_time = datetime.datetime.strptime(lastDate,'%Y/%m/%d')
                except ValueError:  
                    try:
                        middle_time = datetime.datetime.strptime(lastDate,'%Y-%m-%d')
                    except ValueError:  
                        print('历次检查日期格式不正确(10008)')

            # 最后一次检查的时间戳
            lastDate_timeStamp = int(time.mktime(middle_time.timetuple()))
            nextDate_timeStamp = lastDate_timeStamp
            # print(lastDate_timeStamp)

            cell_d_sub = worksheet.cell(row=i,column=cell_d.column)
            risk_level = cell_d_sub.value
            
            if risk_level == 'None' or risk_level == None or len(risk_level) == 0: #值为空 
                prompt = cell_curr.value
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '缺少风险等级'
                else:
                    prompt = prompt + '\n缺少风险等级'
                cell_curr.value = prompt

                cell_curr.fill = PatternFill(patternType='solid',fgColor='FF99FF')
                # cell字体格式设置
                cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('低风险' in risk_level) == True: #只需检查一次
                print('低风险')

                prompt = cell_curr.value
                if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                    prompt = '已完成检查'
                else:
                    prompt = prompt + '\n已完成检查'
                cell_curr.value = prompt


                cell_curr.fill = PatternFill(patternType='solid',fgColor='B9E0A5')
                cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = ''
                

            elif ('一般风险' in risk_level) == True:
                print('一般风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 * 3

                if len(dateArr) < 3:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '检查次数不足3次'
                    else:
                        prompt = prompt + '\n检查次数不足3次'

                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F8CECC')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F8CECC')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('较大风险' in risk_level) == True:
                print('较大风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 * 3

                if len(dateArr) < 4:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '检查次数不足4次'
                    else:
                        prompt = prompt + '\n检查次数不足4次'

                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='F19C99')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")

            elif ('重大风险' in risk_level) == True:
                print('重大风险')
                nextDate_timeStamp = lastDate_timeStamp + 3600 * 24 * 30 

                # 下次检查时间
                nextDate = time.strftime('%Y.%m',time.localtime(nextDate_timeStamp))
                cell_e_sub = worksheet.cell(row=i,column=cell_e.column)
                cell_e_sub.value = nextDate

                # 是否逾期未检查
                t = int(time.time())
                if t > nextDate_timeStamp:
                    prompt = cell_curr.value
                    if prompt == 'None' or prompt == None or len(prompt) == 0: #值为空 
                        prompt = '逾期未检查'
                    else:
                        prompt = prompt + '\n逾期未检查'
                
                    cell_curr.value = prompt

                    cell_curr.fill = PatternFill(patternType='solid',fgColor='EA6B66')
                    cell_curr.font = Font(name="Arial", size=11, color="FF2600", underline="none")


    # 当前时间戳
    t = int(time.time())
    time_difference = t - st_timeStamp
    

    #cell边框设置
    pink = "00FF00FF"
    black = '333333'
    green = "00008000"
    thin = Side(border_style="thin", color=black)
    double = Side(border_style="double", color=green)
    cell_curr.border = Border(top=thin, left=double, right=thin, bottom=thin)







t = time.time()
t1 = time.strftime('%Y-%m-%d',time.localtime(t))
fileName = fileName.replace('.xlsx', '')
fileName = fileName.replace('.xls', '')
outputPath = os.path.dirname(os.getcwd())
outputPath = outputPath + '/' + fileName + ' ' + t1 + '.xlsx'
print('请查看Excel文件:' + outputPath)
myworkbook.save(outputPath)
    

