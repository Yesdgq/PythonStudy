#!/usr/bin/python3

import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

filePath = os.path.abspath(os.path.join(os.getcwd(),".."))
filePath = filePath + '/二组项目信息.xlsx'

print(filePath)
myworkbook = load_workbook(filePath) 
# print(myworkbook.sheetnames) 
worksheet = myworkbook['房建（在施）']
cell = worksheet.cell(row=worksheet['Ao3'].row,column=worksheet['Ao3'].column)
cell.value = '该检查了'
cell.fill = PatternFill(patternType='solid',fgColor='FFFF00')#黄
# cell字体格式设置
cell.font = Font(name="Arial", size=18, color="0057A6", underline="none")
# cell对齐方式
cell.alignment = Alignment(horizontal='center', vertical='center')

t = time.time()
t1 = time.strftime('%Y-%m-%d',time.localtime(t))
outputPath = os.path.dirname(os.getcwd())
outputPath = outputPath + '/输出/二组项目信息 ' + t1 + '.xlsx'
print(outputPath)
myworkbook.save(outputPath)