#!/usr/bin/python3

# import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import os


'''
# 创建Excel
wk = Workbook() #会自动创建一个sheet
# 创建Sheet

# sheet = wk[wk.sheetnames[0]]
# sheet.title = "中国"
# print(wk.sheetnames)

sheet = wk.create_sheet("中国")
sheet = wk.create_sheet("日本")
sheet = wk.create_sheet("美国")

# 在sheet中保存数据
# sheet.append(["aaa", "bbb"])
# wk.save('data222.xlsx') #保存到当前执行文件的路径下
wk.save("/Users/yesdgq/Desktop/excelStudy.xlsx") #保存到指定路径下

'''

path = "/Users/yesdgq/Desktop/excelStudy.xlsx"

myworkbook = load_workbook(path) 
print(myworkbook.sheetnames) 

worksheet = myworkbook[myworkbook.sheetnames[0]]
worksheet.title = "China"
print(myworkbook.sheetnames)

worksheet = myworkbook['中国']
worksheet['A1']='Hello Python' 
worksheet['B5']='新年快乐'

worksheet = myworkbook[myworkbook.sheetnames[1]]
worksheet['B5']='小日本'

# 读取指定cell的值
worksheet = myworkbook[myworkbook.sheetnames[2]]
worksheet['f5']='美国佬🇺🇸'
worksheet['f3'].value='python'

# cell字体格式设置
worksheet['f3'].font = Font(name="Arial", size=18, color="00FF0000", underline="none")

# cell对齐方式
worksheet['f3'].alignment = Alignment(horizontal='right', vertical='top')

# cell对齐方式-旋转角度
worksheet['f3'].alignment = Alignment(text_rotation=90)

#cell边框设置
pink = "00FF00FF"
green = "00008000"
thin = Side(border_style="thin", color=pink)
double = Side(border_style="double", color=green)
worksheet['f3'].border = Border(top=double, left=thin, right=thin, bottom=double)


# sheet表的最大行、最大列
rows = worksheet.max_row #行数
columns = worksheet.max_column #列数
print(rows, columns)

# 读取指定cell的值
cell_value = worksheet.cell(row=5, column=6).value
print(cell_value)



worksheet.cell(row=rows,column=columns).fill = PatternFill(patternType='solid',fgColor='FFFF00')#黄
worksheet.cell(row=rows,column=columns).font = Font(color='0057A6')#红色字体
worksheet.cell(row=rows,column=columns).value = "边界"


# China表
# 颜色填充
worksheet = myworkbook[myworkbook.sheetnames[0]]
for rows in worksheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=6):
    for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color="00FFFF00", end_color="00008000", fill_type = "solid")


worksheet['a8'] = 'id'
worksheet['b8'] = '姓名'
worksheet['c8'] = '性别'
worksheet['d8'] = '年龄'
worksheet['e8'] = '电话'
worksheet['f8'] = '地址'

data = {
    ('1', '张三', '男', '18', '13590098877', '温特莱中心'),
    ('2', '李四', '男', '18', '13590098877', '温特莱中心'),
    ('3', '王五', '男', '18', '13590098877', '温特莱中心'),
    ('4', '王二麻子', '男', '18', '13590098877', '温特莱中心'),
    ('5', '铁蛋', '男', '18', '13590098877', '温特莱中心')
}

# for i in data:
#     worksheet.append(i) #sheet最大行处拼接

# 把数据转化成list或字典
user_info_dic = {}
for rows in worksheet.iter_rows(min_row=9, max_row=13, min_col=1, max_col=6):
    user_info_dic[rows[1].value]=list(i.value for i in rows)



myworkbook.save(path)
print(user_info_dic)



